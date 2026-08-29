"""Append relevant EIPC transmission-facility projects to the US project workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DESTINATION = ROOT / "reference" / "us_reconductoring_projects.xlsx"
DEFAULT_SOURCE = ROOT / "geoinfo" / "us-data" / "recond_gets_files_by_iso" / "eipc-transmission-facilities-2018-2023-roll-up-cases.xlsx"
DEFAULT_GENERATED_DIR = ROOT / "webmap" / "data" / "reconductoring-us"
DEFAULT_SUBSTATIONS = ROOT / "geoinfo" / "us-data" / "Substations.csv"

DIRECT_SHEET_ISO = {
    "SCPSA": "SERTP",
    "SCE&G": "SERTP",
    "TVA": "SERTP",
    "SBA": "SPP",
    "ISONE": "ISO-NE",
    "MISO": "MISO",
    "Entergy": "MISO",
    "NYISO": "NYISO",
    "PJM": "PJM",
    "SPP": "SPP",
}
DESTINATION_SHEETS = {
    "CAISO": "CAISO",
    "ISO-NE": "ISO-NE",
    "MISO": "MISO",
    "NYISO": "NYISO",
    "PJM": "PJM",
    "SPP": "SPP",
    "ERCOT": "ERCOT",
    "WestConnect": "WestConnect",
    "NorthernGrid": "NorthernGrid",
    "SERTP": "SERTP",
}

RELEVANT_TERMS = re.compile(
    r"reconduct|rebuild|upgrade|replace|refurbish|rating|line addition|new line|"
    r"transmission|transformer|substation|capacitor|reactor|terminate|tap",
    re.IGNORECASE,
)

CANONICAL_COLUMNS = [
    "Project Name",
    "SUB_1",
    "SUB_2",
    "ISO/RTO",
    "Utility",
    "Voltage (kV)",
    "Distance (mi)",
    "Status",
    "Planned Year",
    "Description",
    "Source",
]


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized_name(value: Any) -> str:
    text = clean_value(value).upper()
    text = re.sub(r"\b\d+(?:\.\d+)?\s*K?V\b", "", text)
    text = re.sub(r"[^A-Z0-9]+", "", text)
    return text


def row_text(row: dict[str, str]) -> str:
    return " ".join(row.values())


def is_relevant(row: dict[str, str]) -> bool:
    return bool(RELEVANT_TERMS.search(row_text(row)))


def read_source_rows(source: Path) -> list[tuple[str, dict[str, str]]]:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    rows: list[tuple[str, dict[str, str]]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name == "CompleteProjects":
            continue
        sheet = workbook[sheet_name]
        values = sheet.iter_rows(values_only=True)
        raw_headers = next(values, ())
        headers = [clean_header(value) for value in raw_headers]
        headers = [header for header in headers if header]
        if "From Bus Name" not in headers or "To Bus Name" not in headers:
            continue
        for raw_row in values:
            row = {
                headers[index]: clean_value(raw_row[index] if index < len(raw_row) else "")
                for index in range(len(headers))
            }
            if row.get("From Bus Name") or row.get("To Bus Name"):
                rows.append((sheet_name, row))
    return rows


def load_region_geometry(generated_dir: Path, iso_key: str) -> tuple[list[dict[str, Any]], float]:
    path = generated_dir / f"{iso_key}.json"
    if not path.exists():
        return [], 0.0
    payload = json.loads(path.read_text(encoding="utf-8"))
    features = payload.get("regionFeatures") or []
    return features, sum(geometry_area(feature.get("geometry")) for feature in features)


def geometry_area(geometry: dict[str, Any] | None) -> float:
    if not geometry:
        return 0.0

    def ring_area(ring: list[list[float]]) -> float:
        return abs(sum(
            ring[index][0] * ring[(index + 1) % len(ring)][1]
            - ring[(index + 1) % len(ring)][0] * ring[index][1]
            for index in range(len(ring))
        )) / 2

    if geometry.get("type") == "Polygon":
        rings = geometry.get("coordinates") or []
        return sum(ring_area(ring) * (-1 if index else 1) for index, ring in enumerate(rings))
    if geometry.get("type") == "MultiPolygon":
        return sum(geometry_area({"type": "Polygon", "coordinates": polygon}) for polygon in geometry.get("coordinates") or [])
    return 0.0


def point_in_ring(point: tuple[float, float], ring: list[list[float]]) -> bool:
    x, y = point
    inside = False
    for index, current in enumerate(ring):
        previous = ring[index - 1]
        if (current[1] > y) != (previous[1] > y):
            crossing_x = (previous[0] - current[0]) * (y - current[1]) / (previous[1] - current[1]) + current[0]
            if x < crossing_x:
                inside = not inside
    return inside


def point_in_geometry(point: tuple[float, float], geometry: dict[str, Any] | None) -> bool:
    if not geometry:
        return False
    if geometry.get("type") == "Polygon":
        rings = geometry.get("coordinates") or []
        return bool(rings) and point_in_ring(point, rings[0]) and not any(point_in_ring(point, ring) for ring in rings[1:])
    if geometry.get("type") == "MultiPolygon":
        return any(point_in_geometry(point, {"type": "Polygon", "coordinates": polygon}) for polygon in geometry.get("coordinates") or [])
    return False


def load_substations(path: Path) -> dict[str, list[tuple[float, float, str]]]:
    matches: dict[str, list[tuple[float, float, str]]] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            name = normalized_name(row.get("NAME"))
            if not name:
                continue
            try:
                coordinate = (float(row["LONGITUDE"]), float(row["LATITUDE"]), clean_value(row.get("STATE")))
            except (KeyError, TypeError, ValueError):
                continue
            matches.setdefault(name, []).append(coordinate)
    return matches


def route_mapp(row: dict[str, str], substations: dict[str, list[tuple[float, float, str]]], regions: dict[str, tuple[list[dict[str, Any]], float]]) -> str:
    states = set(re.findall(r"[A-Z]{2}", row.get("States", "").upper()))
    endpoint_points = []
    for field in ("From Bus Name", "To Bus Name"):
        candidates = substations.get(normalized_name(row.get(field))) or []
        endpoint_points.append(next((candidate for candidate in candidates if candidate[2] in states), candidates[0] if candidates else None))

    scores = {}
    for iso_key in ("miso", "spp"):
        features = regions[iso_key][0]
        scores[iso_key] = sum(
            1 for candidate in endpoint_points
            if candidate and any(point_in_geometry((candidate[0], candidate[1]), feature.get("geometry")) for feature in features)
        )
    if scores["miso"] != scores["spp"]:
        return "MISO" if scores["miso"] > scores["spp"] else "SPP"
    state_scores = {
        "miso": len(states & {"MT", "ND", "SD", "MN", "IA", "WI", "MI", "IN", "IL", "MO", "KY", "AR", "MS", "LA", "TX"}),
        "spp": len(states & {"AR", "IA", "KS", "LA", "MN", "MO", "MT", "NE", "NM", "ND", "OK", "SD", "TX"}),
    }
    if state_scores["miso"] != state_scores["spp"]:
        return "MISO" if state_scores["miso"] > state_scores["spp"] else "SPP"
    return "MISO" if regions["miso"][1] >= regions["spp"][1] else "SPP"


def to_destination_row(source_sheet: str, row: dict[str, str], iso: str) -> dict[str, str]:
    from_name = row.get("From Bus Name", "")
    to_name = row.get("To Bus Name", "")
    description = row.get("Project Description", "")
    project_type = row.get("Project Type", "")
    label = f"{from_name} - {to_name} {project_type or 'transmission project'}".strip(" -")
    result = {
        "Project Name": label,
        "SUB_1": from_name,
        "SUB_2": to_name,
        "ISO/RTO": iso,
        "Utility": row.get("PA", ""),
        "Voltage (kV)": row.get("Voltage (kV)", ""),
        "Distance (mi)": row.get("Line Length (miles)", ""),
        "Status": row.get("Planning Status", ""),
        "Planned Year": row.get("Model Year", ""),
        "Description": description,
        "Source": f"EIPC roll-up 2018-2023 ({source_sheet})",
    }
    result.update({f"EIPC {key}": value for key, value in row.items() if value})
    return result


def row_key(row: dict[str, str]) -> tuple[str, str, str, str]:
    return (
        normalized_name(row.get("Project Name")),
        normalized_name(row.get("SUB_1")),
        normalized_name(row.get("SUB_2")),
        clean_value(row.get("Description")).lower(),
    )


def merge_workbook(destination: Path, source: Path, generated_dir: Path, substations_path: Path, dry_run: bool) -> None:
    source_rows = read_source_rows(source)
    regions = {
        key: load_region_geometry(generated_dir, key)
        for key in ("miso", "spp")
    }
    substations = load_substations(substations_path)
    additions: dict[str, list[dict[str, str]]] = {sheet: [] for sheet in DESTINATION_SHEETS.values()}
    route_counts: dict[str, int] = {}
    for source_sheet, source_row in source_rows:
        if not is_relevant(source_row):
            continue
        iso = DIRECT_SHEET_ISO.get(source_sheet)
        if source_sheet == "MAPP":
            iso = route_mapp(source_row, substations, regions)
        if not iso:
            continue
        additions[DESTINATION_SHEETS[iso]].append(to_destination_row(source_sheet, source_row, iso))
        route_counts[iso] = route_counts.get(iso, 0) + 1

    workbook = openpyxl.load_workbook(destination)
    total_added = 0
    for sheet_name, rows in additions.items():
        if not rows:
            continue
        sheet = workbook[sheet_name]
        headers = [clean_header(cell.value) for cell in sheet[1]]
        existing_rows = {
            row_key({headers[index]: clean_value(values[index]) for index in range(min(len(headers), len(values)))})
            for values in sheet.iter_rows(min_row=2, values_only=True)
        }
        new_headers = [key for row in rows for key in row if key not in headers]
        for header in dict.fromkeys(new_headers):
            headers.append(header)
            sheet.cell(row=1, column=len(headers), value=header)
        for row in rows:
            key = row_key(row)
            if key in existing_rows:
                continue
            sheet.append([row.get(header, "") for header in headers])
            existing_rows.add(key)
            total_added += 1

    print(f"Relevant source rows: {sum(route_counts.values())}")
    print("Routed rows:", ", ".join(f"{key}={value}" for key, value in sorted(route_counts.items())))
    print(f"Rows appended: {total_added}")
    if not dry_run:
        workbook.save(destination)
        print(f"Updated {destination}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--destination", type=Path, default=DEFAULT_DESTINATION)
    parser.add_argument("--generated-dir", type=Path, default=DEFAULT_GENERATED_DIR)
    parser.add_argument("--substations", type=Path, default=DEFAULT_SUBSTATIONS)
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    merge_workbook(args.destination, args.source, args.generated_dir, args.substations, args.dry_run)
