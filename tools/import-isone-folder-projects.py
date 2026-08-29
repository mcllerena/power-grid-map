"""Import current ISO-NE projects from the ISONE source workbooks."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "geoinfo" / "us-data" / "recond_gets_files_by_iso" / "ISONE"
DESTINATION = ROOT / "reference" / "us_reconductoring_projects.xlsx"
SUBSTATIONS = ROOT / "geoinfo" / "us-data" / "Substations.csv"

QUALIFYING = re.compile(
    r"reconduct|rebuild|upgrade|replace|refurb|structure|transmission|line|"
    r"substation|transformer|breaker|reactor|capacitor|switching|facility|rating",
    re.IGNORECASE,
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def norm(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())


def headers(values: tuple[Any, ...]) -> list[str]:
    output = []
    for value in values:
        name = clean(value)
        output.append(name or f"Column {len(output) + 1}")
    return output


def read_table(path: Path, sheet: str, header_row: int) -> list[dict[str, str]]:
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)[sheet]
    rows = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows, ())
    field_names = headers(next(rows, ()))
    result = []
    for values in rows:
        row = {field_names[i]: clean(values[i] if i < len(values) else "") for i in range(len(field_names))}
        if any(row.values()):
            result.append(row)
    return result


def load_substation_names(path: Path) -> list[str]:
    names = set()
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            value = clean(row.get("NAME"))
            if value:
                names.add(value)
    return sorted(names, key=lambda value: len(norm(value)), reverse=True)


def find_endpoints(text: str, known_names: list[str]) -> tuple[str, str]:
    patterns = [
        r"\bfrom\s+(.{3,80}?)\s+to\s+(.{3,80}?)(?:\.|,|;|$)",
        r"\bbetween\s+(.{3,80}?)\s+and\s+(.{3,80}?)(?:\.|,|;|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return clean(match.group(1)), clean(match.group(2))
    return "", ""


def source_row(source_file: str, source_sheet: str, row: dict[str, str], known_names: list[str], source_id: str) -> dict[str, str] | None:
    description = clean(row.get("Description") or row.get("Project") or row.get("Project Component"))
    major = clean(row.get("Major Project") or row.get("Asset Condition Grouping"))
    driver = clean(row.get("Primary Driver"))
    searchable = " ".join((driver, major, description))
    if not QUALIFYING.search(searchable):
        return None
    sub1, sub2 = find_endpoints(searchable, known_names)
    if not sub1:
        sub1 = f"ISO-NE project {source_id}"
        sub2 = "Regional location"
    project_name = clean(description or major or f"ISO-NE project {source_id}")
    result = {
        "Project Name": project_name,
        "SUB_1": sub1,
        "SUB_2": sub2,
        "ISO/RTO": "ISO-NE",
        "Utility": row.get("Primary Equipment Owner") or row.get("Responsible TO") or "",
        "Voltage (kV)": row.get("Voltage (kV)") or "",
        "Distance (mi)": row.get("Distance (mi)") or "",
        "Status": row.get("Latest Status") or row.get("Planning Status") or "",
        "Planned Year": (row.get("Projected In-Service Year") or row.get("Projected In-Service Month/Year") or row.get("FCM Certified In-Service Date") or "")[:4],
        "Description": description,
        "Source": f"ISONE folder ({source_file} / {source_sheet})",
        "ISONE Source ID": source_id,
        "ISONE Primary Driver": driver,
        "ISONE Major Project": major,
        "ISONE Project Type": row.get("Project Type") or "",
        "ISONE Source State": row.get("State") or "",
    }
    return result


def row_key(row: dict[str, str]) -> tuple[str, str, str, str]:
    return (norm(row.get("ISONE Source ID")), norm(row.get("SUB_1")), norm(row.get("SUB_2")), norm(row.get("Description")))


def load_sources(known_names: list[str]) -> list[dict[str, str]]:
    rows = []
    acl = "final_asset_condition_list_jun_2026.xlsx"
    rsp = "final_rsp_project_list_jun_2026.xlsx"
    fcm = "fcm-certified-transmission-projects-jan-2025.xlsx"
    for filename, sheet, header, id_field in [
        (acl, "06_2026_ACL", 2, "Asset Condition ID"),
        (rsp, "06_2026_RSP", 2, "Project ID"),
    ]:
        for row in read_table(SOURCE_DIR / filename, sheet, header):
            item = source_row(filename, sheet, row, known_names, f"{sheet}-{row.get(id_field) or 'row'}")
            if item:
                rows.append(item)
    for index, row in enumerate(read_table(SOURCE_DIR / fcm, "Currently Certified Projects", 10), 1):
        item = source_row(fcm, "Currently Certified Projects", row, known_names, row.get("RSP Project ID / AC Project ID / LSP") or f"FCM-{index}")
        if item:
            rows.append(item)
    return rows


def merge(destination: Path, dry_run: bool) -> None:
    known_names = load_substation_names(SUBSTATIONS)
    candidates = load_sources(known_names)
    workbook = openpyxl.load_workbook(destination)
    sheet = workbook["ISO-NE"]
    existing_headers = [clean(cell.value) for cell in sheet[1]]
    existing = set()
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {existing_headers[i]: clean(values[i] if i < len(values) else "") for i in range(len(existing_headers))}
        existing.add(row_key(row))
    for row in candidates:
        for key in row:
            if key not in existing_headers:
                existing_headers.append(key)
                sheet.cell(row=1, column=len(existing_headers), value=key)
    added = 0
    seen = set(existing)
    for row in candidates:
        key = row_key(row)
        if key in seen:
            continue
        sheet.append([row.get(header, "") for header in existing_headers])
        seen.add(key)
        added += 1
    print(f"Qualifying ISONE rows: {len(candidates)}")
    print(f"Unique rows appended: {added}")
    print(f"Rows skipped as duplicates: {len(candidates) - added}")
    if not dry_run:
        workbook.save(destination)
        print(f"Updated {destination}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--destination", type=Path, default=DESTINATION)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    merge(args.destination, args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
